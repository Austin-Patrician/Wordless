#!/usr/bin/env python3
"""Inspect common local tabular files without loading unbounded data."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", type=Path)
    parser.add_argument("--sample-rows", type=int, default=20)
    parser.add_argument("--encoding", default=None)
    parser.add_argument("--max-json-bytes", type=int, default=200_000_000)
    parser.add_argument("--output", type=Path, default=None)
    return parser.parse_args()


def detect_encoding(path: Path, encoding: str | None) -> str:
    candidates = [encoding] if encoding else ["utf-8-sig", "gb18030", "utf-16"]
    sample = path.read_bytes()[:1_048_576]
    for candidate in candidates:
        if candidate is None:
            continue
        try:
            sample.decode(candidate)
            return candidate
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode text as UTF-8, GB18030, or UTF-16; pass --encoding explicitly")


def read_text(path: Path, encoding: str | None) -> tuple[str, str]:
    used_encoding = detect_encoding(path, encoding)
    return path.read_text(encoding=used_encoding), used_encoding


def scalar_type(value: Any) -> str:
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, (dict, list)):
        return "object"
    text = str(value).strip()
    if not text:
        return "null"
    try:
        int(text)
        return "integer"
    except ValueError:
        try:
            float(text)
            return "number"
        except ValueError:
            return "string"


def column_info(headers: list[str], rows: Iterable[list[Any]]) -> tuple[list[dict[str, Any]], int, int, list[dict[str, Any]]]:
    types = [set[str]() for _ in headers]
    nulls = [0 for _ in headers]
    row_count = 0
    malformed = 0
    samples: list[dict[str, Any]] = []
    for row in rows:
        row_count += 1
        if len(row) != len(headers):
            malformed += 1
        values = row[: len(headers)] + [None] * max(0, len(headers) - len(row))
        for index, value in enumerate(values[: len(headers)]):
            value_type = scalar_type(value)
            types[index].add(value_type)
            if value_type == "null":
                nulls[index] += 1
        if len(samples) < 20:
            samples.append({headers[index]: values[index] for index in range(len(headers))})
    columns = []
    for index, name in enumerate(headers):
        observed = sorted(types[index] - {"null"})
        inferred = observed[0] if len(observed) == 1 else "mixed" if observed else "null"
        columns.append({"name": name, "inferredType": inferred, "nullCount": nulls[index]})
    return columns, row_count, malformed, samples


def inspect_delimited(path: Path, sample_rows: int, encoding: str | None) -> dict[str, Any]:
    used_encoding = detect_encoding(path, encoding)
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding=used_encoding, newline="") as handle:
        preview = "".join(handle.readline() for _ in range(20))
        try:
            delimiter = csv.Sniffer().sniff(preview, delimiters=",\t;|").delimiter
        except csv.Error:
            pass
        handle.seek(0)
        reader = csv.reader(handle, delimiter=delimiter)
        header_row = next(reader, [])
        headers = [value.strip() or f"column_{index + 1}" for index, value in enumerate(header_row)]
        columns, row_count, malformed, samples = column_info(headers, reader)
    return {
        "format": "tsv" if path.suffix.lower() == ".tsv" else "csv",
        "encoding": used_encoding,
        "delimiter": delimiter,
        "datasets": [{"name": path.stem, "rows": row_count, "columns": columns, "sample": samples[:sample_rows]}],
        "warnings": [f"{malformed} rows have a different number of fields"] if malformed else [],
    }


def inspect_json(path: Path, sample_rows: int, max_bytes: int, encoding: str | None) -> dict[str, Any]:
    if path.suffix.lower() == ".jsonl":
        return inspect_jsonl(path, sample_rows, encoding)
    if path.stat().st_size > max_bytes:
        raise ValueError(f"JSON input exceeds --max-json-bytes ({max_bytes}) and cannot be safely loaded")
    text, used_encoding = read_text(path, encoding)
    value = json.loads(text)
    records = value if isinstance(value, list) else [value]
    if not records:
        return {"format": path.suffix.lower().lstrip("."), "encoding": used_encoding, "datasets": [{"name": path.stem, "rows": 0, "columns": [], "sample": []}], "warnings": []}
    if not all(isinstance(record, dict) for record in records):
        return {"format": path.suffix.lower().lstrip("."), "encoding": used_encoding, "datasets": [{"name": path.stem, "rows": len(records), "columns": [{"name": "value", "inferredType": scalar_type(records[0]), "nullCount": 0}], "sample": [{"value": record} for record in records[:sample_rows]]}], "warnings": []}
    headers = list(dict.fromkeys(key for record in records for key in record))
    columns, row_count, malformed, samples = column_info(headers, ([record.get(key) for key in headers] for record in records))
    return {"format": path.suffix.lower().lstrip("."), "encoding": used_encoding, "datasets": [{"name": path.stem, "rows": row_count, "columns": columns, "sample": samples[:sample_rows]}], "warnings": [f"{malformed} records have missing fields"] if malformed else []}


def inspect_jsonl(path: Path, sample_rows: int, encoding: str | None) -> dict[str, Any]:
    used_encoding = detect_encoding(path, encoding)
    headers: list[str] = []
    types: list[set[str]] = []
    nulls: list[int] = []
    samples: list[dict[str, Any]] = []
    row_count = 0
    malformed = 0
    with path.open("r", encoding=used_encoding) as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            value = json.loads(line)
            row_count += 1
            if not isinstance(value, dict):
                if not headers:
                    headers = ["value"]
                    types = [set[str]()]
                    nulls = [0]
                types[0].add(scalar_type(value))
                if len(samples) < sample_rows:
                    samples.append({"value": value})
                continue
            if any(not isinstance(key, str) for key in value):
                malformed += 1
                continue
            for key in value:
                if key not in headers:
                    headers.append(key)
                    types.append(set[str]())
                    nulls.append(0)
            for index, key in enumerate(headers):
                value_type = scalar_type(value.get(key))
                types[index].add(value_type)
                if value_type == "null":
                    nulls[index] += 1
            if len(samples) < sample_rows:
                samples.append({key: value.get(key) for key in headers})
    columns = []
    for index, name in enumerate(headers):
        observed = sorted(types[index] - {"null"})
        inferred = observed[0] if len(observed) == 1 else "mixed" if observed else "null"
        columns.append({"name": name, "inferredType": inferred, "nullCount": nulls[index]})
    warnings = [f"{malformed} records were not objects"] if malformed else []
    return {"format": "jsonl", "encoding": used_encoding, "datasets": [{"name": path.stem, "rows": row_count, "columns": columns, "sample": samples}], "warnings": warnings}


def inspect_xlsx(path: Path, sample_rows: int) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError("openpyxl is required to inspect XLSX files") from error
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    datasets = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            header_row = list(next(iterator, ()))
            headers = [str(value).strip() if value not in (None, "") else f"column_{index + 1}" for index, value in enumerate(header_row)]
            columns, row_count, malformed, samples = column_info(headers, (list(row) for row in iterator))
            datasets.append({"name": sheet_name, "rows": row_count, "columns": columns, "sample": samples[:sample_rows], "warnings": [f"{malformed} rows have a different number of cells"] if malformed else []})
    finally:
        workbook.close()
    return {"format": "xlsx", "datasets": datasets, "warnings": []}


def inspect_parquet(path: Path, sample_rows: int) -> dict[str, Any]:
    try:
        import pyarrow.parquet as parquet
    except ImportError as error:
        raise RuntimeError("pyarrow is required to inspect Parquet files") from error
    file = parquet.ParquetFile(path)
    schema = file.schema_arrow
    columns = [{"name": field.name, "inferredType": str(field.type), "nullCount": None} for field in schema]
    return {"format": "parquet", "datasets": [{"name": path.stem, "rows": file.metadata.num_rows, "columns": columns, "sample": [], "sampleRowsRequested": sample_rows}], "warnings": ["Parquet schema metadata was inspected without loading row data"]}


def main() -> int:
    args = parse_args()
    if args.sample_rows < 0:
        raise ValueError("--sample-rows must be non-negative")
    path = args.path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        result = inspect_delimited(path, args.sample_rows, args.encoding)
    elif suffix in {".json", ".jsonl"}:
        result = inspect_json(path, args.sample_rows, args.max_json_bytes, args.encoding)
    elif suffix in {".xlsx", ".xlsm", ".xltx"}:
        result = inspect_xlsx(path, args.sample_rows)
    elif suffix == ".parquet":
        result = inspect_parquet(path, args.sample_rows)
    else:
        raise ValueError(f"Unsupported data format: {suffix or '(none)'}")
    result.update({"path": str(path), "fileSizeBytes": path.stat().st_size, "inspectedAt": datetime.now().astimezone().isoformat()})
    encoded = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(encoded + "\n", encoding="utf-8")
    print(encoded)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(json.dumps({"error": str(error)}, ensure_ascii=False), file=sys.stderr)
        raise SystemExit(2)
